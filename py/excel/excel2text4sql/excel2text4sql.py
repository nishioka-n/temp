# 翻訳リストのExcelファイルからSQL(DML: DELETE, INSERT)を生成しファイル出力するツール
# 
# PEP8 には従っていません。（社内ツールなので、ご容赦ください...）
# 

import os
import sys
import openpyxl as px

# 各種設定定義
from config import *


# メッセージ用エスケープ処理（翻訳リストのExcel関数より）
def escape4sql(value):
    escaped = value \
        .replace('"', '\\"').replace("'", "\\'") \
        .replace("(", "\\(").replace(")", "\\)")
    if ESCAPE_LF:        
        escaped = escaped.replace("\n", "\\n")  # 既存に合わせて、 \r\n に対応する場合は要改修
    return escaped


# Excelの列名を列番号へ変換（A → 1）
def get_col_index(col_name):
    return px.utils.cell.column_index_from_string(col_name)


# テキストファイル出力
def write_to_text_file(txt_file_name, lines, encoding="UTF-8", newline="\n"):
    with open(txt_file_name, "w", encoding=encoding, newline=newline) as f:
        for line in lines:
            f.write(line + newline)


# 各シートごとの処理
def process_worksheet(ws):
    max_row = ws.max_row
    # print(f"sheet='{ws.title}' max_row={max_row}")

    del_lines = []
    ins_lines = []

    # 共通設定
    language_count = common_config['language_count']

    # シート別設定
    cs = sheet_config[ws.title]
    language_code_row = cs['language_code_row'] 
    item_start_row = cs['item_start_row']
    item_type_col = get_col_index(cs['item_type_col'])
    item_id_col = get_col_index(cs['item_id_col'])
    language_code_1st_col = get_col_index(cs['language_code_1st_col'])
    target_cond_col = get_col_index(cs['target_cond_col'])
    target_cond_str = cs['target_cond_str']
    target_language_codes = cs['target_language_codes']

    # 全言語コードが対象かどうか
    is_all_lang = (len(target_language_codes) == 0 or language_count == len(target_language_codes))

    min_row = item_start_row
    max_col = language_code_1st_col + language_count

    # 言語コードを取得
    lang_code_list = []
    for col in range(language_code_1st_col, max_col):
        lang_code_list.append(ws.cell(column=col, row=language_code_row).value)

    # 各行の処理（速度の問題があり、イテレータで各行を取得し、リストインデックスでのセル参照方式採用）
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, max_col=max_col):
        # 条件が指定されていれば、一致しない行を除外
        if target_cond_str:
            condition_value = row[target_cond_col - 1].value
            if condition_value != target_cond_str:
                continue

        item_type = row[item_type_col - 1].value
        item_id = row[item_id_col - 1].value
        if not item_type:
            continue
        print(f"{item_type}-{item_id}")

        # 各列の処理（言語数分）
        for col_index in range(language_code_1st_col, max_col):
            language_code = lang_code_list[col_index - language_code_1st_col]
            # 言語コードがすべて対象でなければ、指定の言語コード以外は除外する
            if not is_all_lang and not language_code in target_language_codes:
                continue

            message_cell = row[col_index - 1]
            message = message_cell.value
            if not message:
                continue

            # メッセージのエスケープ処理
            message = escape4sql(message)
            # DELETE文生成
            if not is_all_lang:
                # 言語コードごと
                del_sql = DELETE_SQL_LANG.format(item_type=item_type, item_id=item_id, language_code=language_code)
                del_lines.append(del_sql)
            elif col_index == language_code_1st_col:
                # アイテムごと（1回だけ生成）： メッセージの有無まで見ないと不要な出力してしまうので、仕方なくこの位置にある。
                del_sql = DELETE_SQL_ITEM.format(item_type=item_type, item_id=item_id)
                del_lines.append(del_sql)
            # INSERT文生成
            ins_sql = INSERT_SQL.format(item_type=item_type, item_id=item_id, language_code=language_code, message=message)
            ins_lines.append(ins_sql)

    return del_lines, ins_lines


# ワークブックの処理
def process_workbook(xlsx_file_name):

    # 対象シート名
    sheet_names = common_config['sheet_names']

    wb = px.load_workbook(xlsx_file_name, data_only=True, read_only=True)
    lines = []
    for ws in wb.worksheets:
        if not ws.title in sheet_names:
            continue
        lines.append("")
        lines.append(f"-- {ws.title}")

        # 各シートの処理
        del_lines, ins_lines = process_worksheet(ws)

        lines.extend(del_lines)
        lines.extend(ins_lines)
    return lines


def main(args):

    # 入力ファイル指定（一時的に違うファイルを指定する場合は、コマンドライン第一引数で指定）
    input_file_name = args[0] if len(args) > 0 else common_config['input_file_name']
    print(f"入力ファイル: {input_file_name}")
    if not os.path.isfile(input_file_name):
        print("ファイルが見つかりません。")
        sys.exit(1)

    # SQL生成処理
    lines = process_workbook(input_file_name)
    # [print(line) for line in lines]

    # 出力ファイル指定（一時的に違うファイルを指定する場合は、コマンドライン第二引数で指定）
    output_file_name = args[1] if len(args) > 1 else common_config['output_file_name']
    print(f"出力ファイル: {output_file_name}")

    # ファイル出力
    output_encoding = common_config['output_encoding']
    output_newline = common_config['output_newline']
    write_to_text_file(output_file_name, lines, output_encoding, output_newline)

    print("処理完了")

if __name__ == '__main__':
    args = sys.argv[1:]
    main(args)
