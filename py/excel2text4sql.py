import os
import sys
import openpyxl as px
from glob import glob
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string, get_column_letter

# xy = coordinate_from_string('A4') # returns ('A',4)
# col = column_index_from_string(xy[0]) # returns 1
# row = xy[1]

# 設定項目 ---------- config.py へ切り出し

# ファイル名
XLSX_FILE_NAME = 'サンプル.xlsx'

# 対象シート名
SHEETS_NAMES = ["Display+Pop up", "Emailメッセージ"]

# 改行コードをエスケープするか
ESCAPE_LF = True

# DELETE文のフォーマット（既存の手動作成の書式）
DELETE_SQL_FORMAT = "delete from item_master where item_type='{item_type}' and item_id='{item_id}' and language_code='{language_code}';"

# INSERT文のフォーマット（既存の手動作成の書式）
INSERT_SQL_FORMAT = "insert into item_master values('{item_type}','{item_id}','{language_code}',E'{message}',now());"



# 共通設定
conf_common = {
    # 言語数
    'language_count': 20,
}


# シートごとの設定
conf_sheets = {
    "Display+Pop up": {
        # 言語コードの行
        "language_code_row": 14,
        # リストの開始行
        "item_start_row": 16,
        # 
        "item_type_col": "F",
        "item_id_col": "G",
        "language_code_start_col": "H",
        # 対象を絞り込む条件列
        "target_cond_col": "A", 
         # 対象を絞り込む条件文字列（空文字なら何もしない）
        "target_cond_str": "",
    },
    "Emailメッセージ": {
        "language_code_row": 5,
        "item_start_row": 7,
        "item_type_col": "F",
        "item_id_col": "G",
        "language_code_start_col": "H",
        "target_cond_col": "A",  # 対象を絞り込む条件列
        "target_cond_str": "", # 対象を絞り込む条件文字列（空文字なら何もしない）
    }
}



# 設定項目 ---------- config.py へ切り出し


# SQLのためのエスケープ処理
def escape4sql(value):
    escaped = value.replace('"', '\\"').replace("'", "\\'") \
        .replace("(", "\\(") .replace(")", "\\)")
    if ESCAPE_LF:
        escaped = escaped.replace("\n", "\\n")
    return escaped


# シートごとの処理
def get_worksheet_cell_values(ws):
    max_row = ws.max_row
    print(f"■ sheet='{ws.title}' max_row={max_row}")

    del_lines = []
    ins_lines = []

    language_count = conf_common['language_count'] # 共通設定から

    conf = conf_sheets[ws.title]
    item_start_row = conf['item_start_row']      # シート別設定から
    language_code_row = conf['language_code_row']  # シート別設定から
    item_type_col = conf['item_type_col']  # シート別設定から
    item_id_col = conf['item_id_col']  # シート別設定から
    language_code_start_col = conf['language_code_start_col'] # シート別設定から

    target_cond_col = conf['target_cond_col']
    target_cond_str = conf['target_cond_str']

    item_start_col_index = column_index_from_string(language_code_start_col)

    # 各行の処理
    min_row = item_start_row
    min_col = item_start_col_index
    max_col = item_start_col_index + language_count
    for row_index in range(item_start_row, max_row):  # 
    # for rows in sheet.iter_rows(min_row=min_row, max_col=max_col, max_row=max_row):
    #     data = rows[0].value
        
        if target_cond_str:
            condition_value = ws[target_cond_col + str(row_index)].value
            if condition_value != target_cond_str:
                continue

        item_type = ws[item_type_col + str(row_index)].value
        item_id = ws[item_id_col + str(row_index)].value
        if item_type == None:
            continue
        print(f"{item_type}-{item_id}")

        # 各列の処理（言語数分）
        for col_index in range(item_start_col_index, item_start_col_index + language_count):
            language_code = ws.cell(column=col_index, row=language_code_row).value  # 毎回取得は効率悪いので改善
            cell = ws.cell(column=col_index, row=row_index)
            if not cell.value:
                continue
            # エスケープ処理
            message = escape4sql(cell.value)

            cell_addr = cell.coordinate
            # print(f"{cell_addr}:  '{item_type}-{item_id}', '{language_code}', '{message}'")
            del_sql = DELETE_SQL_FORMAT.format(item_type=item_type, item_id=item_id, language_code=language_code)
            ins_sql = INSERT_SQL_FORMAT.format(item_type=item_type, item_id=item_id, language_code=language_code, message=message)
            del_lines.append(del_sql)
            ins_lines.append(ins_sql)

    return del_lines, ins_lines

# 
def get_workbook_cell_values(xlsx_file_name):
    wb = px.load_workbook(xlsx_file_name, data_only=True, read_only=True)
    lines = []
    for ws in wb.worksheets:
        if not ws.title in SHEETS_NAMES:
            continue
        lines.append(f"-- {ws.title}")
        del_lines, ins_lines = get_worksheet_cell_values(ws)
        lines.extend(del_lines)
        lines.extend(ins_lines)
    return lines


def write_to_text_file(txt_file_name, lines):
    with open(txt_file_name, "w", encoding="UTF-8") as f:
        for line in lines:
            f.write(line + '\n')


def main(args):
    xlsx_file_name = XLSX_FILE_NAME
    lines = get_workbook_cell_values(xlsx_file_name)
    # for line in lines:
    #     print(line)
    write_to_text_file("result.sql", lines)

if __name__ == '__main__':
    args = sys.argv[1:]
    main(args)
